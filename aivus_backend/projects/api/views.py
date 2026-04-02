"""API views for projects app."""

import json
import logging
import uuid as uuid_module
from collections import defaultdict
from datetime import UTC
from datetime import datetime
from decimal import Decimal

import openpyxl
from django.db import IntegrityError
from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

try:
    from django_ratelimit.decorators import ratelimit
except ImportError:
    from django.conf import settings as django_settings

    if not django_settings.DEBUG:
        msg = (
            "django-ratelimit is required in production but not installed. "
            "Run: pip install django-ratelimit"
        )
        raise ImportError(msg) from None

    # Fallback: no-op decorator only in DEBUG mode
    def ratelimit(**kwargs):
        def decorator(func):
            return func

        return decorator


import contextlib

from django.utils import timezone as tz

from aivus_backend.catalog.models import Category
from aivus_backend.catalog.models import Entry
from aivus_backend.catalog.models import Unit
from aivus_backend.core.decorators import public_endpoint
from aivus_backend.core.decorators import require_groups
from aivus_backend.core.enums import BriefStatus
from aivus_backend.core.enums import OfferSource
from aivus_backend.core.enums import OfferStatus
from aivus_backend.core.enums import ProjectStatus
from aivus_backend.projects.ai_brief import analyze_brief
from aivus_backend.projects.ai_brief import analyze_comparison
from aivus_backend.projects.ai_brief import process_chat_message
from aivus_backend.projects.api.serializers import serialize_brief
from aivus_backend.projects.api.serializers import serialize_brief_detail
from aivus_backend.projects.api.serializers import serialize_brief_offer
from aivus_backend.projects.api.serializers import serialize_brief_with_offers
from aivus_backend.projects.api.serializers import serialize_offer
from aivus_backend.projects.api.serializers import serialize_offer_for_client
from aivus_backend.projects.api.serializers import serialize_project
from aivus_backend.projects.api.serializers import serialize_rate_card
from aivus_backend.projects.api.serializers import serialize_rate_card_item
from aivus_backend.projects.api.serializers import serialize_share
from aivus_backend.projects.api.serializers import serialize_share_public
from aivus_backend.projects.api.serializers import serialize_template
from aivus_backend.projects.models import Brief
from aivus_backend.projects.models import BriefOffer
from aivus_backend.projects.models import ChatMessage
from aivus_backend.projects.models import ClientManager
from aivus_backend.projects.models import Offer
from aivus_backend.projects.models import OfferDeliverable
from aivus_backend.projects.models import OfferEntry
from aivus_backend.projects.models import OfferScheduleEntry
from aivus_backend.projects.models import Project
from aivus_backend.projects.models import ProjectCollaborator
from aivus_backend.projects.models import RateCard
from aivus_backend.projects.models import RateCardItem
from aivus_backend.projects.models import Share
from aivus_backend.projects.models import Template
from aivus_backend.projects.services import parse_offer_details_to_entries
from aivus_backend.projects.services import recalculate_offer_totals
from aivus_backend.projects.services import reconstruct_details_from_entries
from aivus_backend.users.models import Client
from aivus_backend.users.models import Team
from aivus_backend.users.models import User
from aivus_backend.users.models import Vendor
from aivus_backend.users.models import VendorSettings

logger = logging.getLogger(__name__)

MAX_REQUEST_BODY_SIZE = 1_000_000


def _validate_uuid(value, field_name="id"):
    """Validate and return a UUID string. Raises ValueError if invalid."""
    try:
        uuid_module.UUID(str(value))
        return str(value)
    except (ValueError, AttributeError) as exc:
        msg = f"Invalid UUID format for {field_name}"
        raise ValueError(msg) from exc


# ==================== Projects API ====================


@csrf_exempt
@require_http_methods(["GET", "POST"])
@require_groups("VENDOR", "CLIENT", "SYSTEM")
def projects_list(request):  # noqa: C901, PLR0912, PLR0915
    """List all projects or create a new one."""
    if request.method == "GET":
        # QA3-009: Use authenticated user_data instead of raw header
        vendor_id = request.user_data.get("vendor_id")
        if not vendor_id:
            return JsonResponse({"error": "Vendor ID required"}, status=400)

        projects = (
            Project.objects.filter(
                vendor_id=vendor_id,
            )
            .select_related("client")
            .prefetch_related("collaborators", "client_managers")
        )
        return JsonResponse([serialize_project(p) for p in projects], safe=False)

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            vendor_id = data.get("vendorId")
            brief_id = data.get("briefId")
            team_id = data.get("teamId")  # Optional
            name = data.get("name")
            status = data.get("status", "DRAFT")

            # QA4-022: Validate status against ProjectStatus enum
            valid_statuses = [s.value for s in ProjectStatus]
            if status not in valid_statuses:
                return JsonResponse(
                    {
                        "error": (
                            f"Invalid status. Must be one of:"
                            f" {', '.join(valid_statuses)}"
                        )
                    },
                    status=400,
                )

            # New fields
            crm_id = data.get("crmId", "")
            description = data.get("description", "")
            client_id = data.get("clientId")
            client_name = data.get("clientName", "")
            irs_ein = data.get("irsEin", "")
            brand_name = data.get("brandName", "")
            agency_name = data.get("agencyName", "")
            collaborators = data.get("collaborators", [])
            client_managers = data.get("clientManagers", [])

            if not vendor_id or not name:
                return JsonResponse(
                    {"error": "vendorId and name are required"},
                    status=400,
                )

            # QA3-010: Verify vendorId matches authenticated user's vendor
            user_vendor_id = request.user_data.get("vendor_id")
            if not user_vendor_id or vendor_id != user_vendor_id:
                return JsonResponse({"error": "Access denied"}, status=403)

            # Validate UUID fields
            try:
                _validate_uuid(vendor_id, "vendorId")
                if brief_id:
                    _validate_uuid(brief_id, "briefId")
                if team_id:
                    _validate_uuid(team_id, "teamId")
                if client_id:
                    _validate_uuid(client_id, "clientId")
            except ValueError as e:
                return JsonResponse({"error": str(e)}, status=400)

            # Verify vendor exists
            try:
                vendor = Vendor.objects.get(id=vendor_id)
            except Vendor.DoesNotExist:
                return JsonResponse({"error": "Vendor not found"}, status=404)

            # Verify team exists if provided (non-fatal if not found)
            team = None
            if team_id:
                try:
                    team = Team.objects.get(id=team_id)
                except Team.DoesNotExist:
                    logger.warning(
                        "Team %s not found, creating project without team", team_id
                    )

            # Verify brief exists if provided
            brief = None
            if brief_id:
                try:
                    brief = Brief.objects.get(id=brief_id)
                except Brief.DoesNotExist:
                    return JsonResponse({"error": "Brief not found"}, status=404)

            # Verify client exists if provided
            client = None
            if client_id:
                try:
                    client = Client.objects.get(id=client_id)
                except Client.DoesNotExist:
                    return JsonResponse({"error": "Client not found"}, status=404)

            project = Project.objects.create(
                name=name,
                vendor=vendor,
                brief=brief,
                team=team,
                status=status,
                crm_id=crm_id,
                description=description,
                client=client,
                client_name=client_name,
                irs_ein=irs_ein,
                brand_name=brand_name,
                agency_name=agency_name,
            )

            # Create collaborators
            for collab in collaborators:
                user = None
                user_id = collab.get("userId")
                if user_id:
                    with contextlib.suppress(User.DoesNotExist):
                        user = User.objects.get(id=user_id)

                ProjectCollaborator.objects.create(
                    project=project,
                    user=user,
                    name=collab.get("name", ""),
                    email=collab.get("email", ""),
                    role=collab.get("role", "internal_user"),
                )

            # Create client managers
            for manager in client_managers:
                ClientManager.objects.create(
                    project=project,
                    name=manager.get("name", ""),
                    position=manager.get("position", ""),
                )

            return JsonResponse(serialize_project(project), status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error creating project")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET"])
@require_groups("VENDOR", "SYSTEM")
def projects_archived(request):
    """List all archived (soft-deleted) projects for the authenticated vendor."""
    vendor_id = request.user_data.get("vendor_id")
    if not vendor_id:
        return JsonResponse({"error": "Vendor ID required"}, status=400)

    projects = (
        Project.objects.all_with_deleted()
        .filter(
            vendor_id=vendor_id,
            deleted_at__isnull=False,
        )
        .select_related("client")
        .prefetch_related("collaborators", "client_managers")
    )
    return JsonResponse([serialize_project(p) for p in projects], safe=False)


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("VENDOR", "SYSTEM")
def project_restore(request, project_id):
    """Restore an archived project."""
    vendor_id = request.user_data.get("vendor_id")
    if not vendor_id:
        return JsonResponse({"error": "Vendor ID required"}, status=400)
    try:
        project = Project.objects.all_with_deleted().get(
            id=project_id,
            vendor_id=vendor_id,
            deleted_at__isnull=False,
        )
    except Project.DoesNotExist:
        return JsonResponse({"error": "Archived project not found"}, status=404)
    project.restore()
    return JsonResponse(serialize_project(project))


@csrf_exempt
@require_http_methods(["GET", "PUT", "PATCH", "DELETE"])
@require_groups("VENDOR", "CLIENT", "SYSTEM")
def project_detail(request, project_id):  # noqa: C901, PLR0912, PLR0915
    """Get, update, or delete a specific project."""
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return JsonResponse({"error": "Project not found"}, status=404)

    # Verify ownership: project's vendor_id must match requesting user's vendor_id
    user_vendor_id = request.user_data.get("vendor_id")
    if not user_vendor_id or str(project.vendor_id) != user_vendor_id:
        return JsonResponse({"error": "Access denied"}, status=403)

    if request.method == "GET":
        return JsonResponse(serialize_project(project))

    if request.method in ["PUT", "PATCH"]:
        try:
            data = json.loads(request.body)

            if "name" in data:
                project.name = data["name"]
            if "status" in data:
                # QA3-034: Validate status against ProjectStatus enum
                valid_statuses = [s.value for s in ProjectStatus]
                if data["status"] not in valid_statuses:
                    return JsonResponse(
                        {
                            "error": (
                                f"Invalid status. Must be one of:"
                                f" {', '.join(valid_statuses)}"
                            )
                        },
                        status=400,
                    )
                project.status = data["status"]
            if "briefId" in data:
                if data["briefId"]:
                    try:
                        brief = Brief.objects.get(id=data["briefId"])
                        project.brief = brief
                    except Brief.DoesNotExist:
                        return JsonResponse({"error": "Brief not found"}, status=404)
                else:
                    project.brief = None
            if "teamId" in data:
                if data["teamId"]:
                    try:
                        team = Team.objects.get(id=data["teamId"])
                        project.team = team
                    except Team.DoesNotExist:
                        return JsonResponse({"error": "Team not found"}, status=404)
                else:
                    project.team = None

            # New fields
            if "crmId" in data:
                project.crm_id = data["crmId"]
            if "description" in data:
                project.description = data["description"]
            if "clientId" in data:
                if data["clientId"]:
                    try:
                        client = Client.objects.get(id=data["clientId"])
                        project.client = client
                    except Client.DoesNotExist:
                        return JsonResponse({"error": "Client not found"}, status=404)
                else:
                    project.client = None
            if "clientName" in data:
                project.client_name = data["clientName"]
            if "irsEin" in data:
                project.irs_ein = data["irsEin"]
            if "brandName" in data:
                project.brand_name = data["brandName"]
            if "agencyName" in data:
                project.agency_name = data["agencyName"]

            # Update collaborators if provided
            if "collaborators" in data:
                with transaction.atomic():
                    # Delete existing and recreate
                    project.collaborators.all().delete()
                    for collab in data["collaborators"]:
                        user = None
                        user_id = collab.get("userId")
                        if user_id:
                            with contextlib.suppress(User.DoesNotExist):
                                user = User.objects.get(id=user_id)

                        ProjectCollaborator.objects.create(
                            project=project,
                            user=user,
                            name=collab.get("name", ""),
                            email=collab.get("email", ""),
                            role=collab.get("role", "internal_user"),
                        )

            # Update client managers if provided
            if "clientManagers" in data:
                with transaction.atomic():
                    # Delete existing and recreate
                    project.client_managers.all().delete()
                    for manager in data["clientManagers"]:
                        ClientManager.objects.create(
                            project=project,
                            name=manager.get("name", ""),
                            position=manager.get("position", ""),
                        )

            project.save()
            return JsonResponse(serialize_project(project))

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error updating project")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    if request.method == "DELETE":
        project.deleted_at = datetime.now(UTC)
        project.save()
        return JsonResponse({"message": "Project deleted"}, status=200)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("VENDOR", "SYSTEM")
def project_thumbnail(request, project_id):
    """Upload project thumbnail image."""
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return JsonResponse({"error": "Project not found"}, status=404)

    # QA3-003: Verify ownership
    user_vendor_id = request.user_data.get("vendor_id")
    if not user_vendor_id or str(project.vendor_id) != user_vendor_id:
        return JsonResponse({"error": "Access denied"}, status=403)

    if "thumbnail" not in request.FILES:
        return JsonResponse({"error": "No file provided"}, status=400)

    file = request.FILES["thumbnail"]

    # Validate file type
    allowed_types = ("image/jpeg", "image/png", "image/gif", "image/webp")
    if file.content_type not in allowed_types:
        return JsonResponse(
            {"error": f"Invalid file type. Allowed types: {', '.join(allowed_types)}"},
            status=400,
        )

    # Validate file size (max 5MB)
    if file.size > 5 * 1024 * 1024:
        return JsonResponse({"error": "File size must not exceed 5MB"}, status=400)

    project.thumbnail = file
    project.save()

    return JsonResponse(
        {
            "thumbnailUrl": project.thumbnail.url if project.thumbnail else None,
        }
    )


# ==================== Briefs API ====================


@csrf_exempt
@require_http_methods(["GET", "POST"])
@require_groups("VENDOR", "CLIENT", "SYSTEM")
def briefs_list(request):
    """List all briefs or create a new one."""
    if request.method == "GET":
        user_group = request.user_data.get("group")
        user_client_id = request.user_data.get("client_id")
        user_vendor_id = request.user_data.get("vendor_id")

        if user_group == "CLIENT" and user_client_id:
            briefs = Brief.objects.filter(
                client_id=user_client_id, deleted_at__isnull=True
            )
        elif user_group == "VENDOR" and user_vendor_id:
            briefs = Brief.objects.filter(
                projects__vendor_id=user_vendor_id,
                deleted_at__isnull=True,
            ).distinct()
        elif user_group == "SYSTEM":
            briefs = Brief.objects.filter(deleted_at__isnull=True)
        else:
            briefs = Brief.objects.none()

        return JsonResponse([serialize_brief(b) for b in briefs], safe=False)

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            status = data.get("status", "DRAFT")
            details = data.get("details", {})
            client_id = data.get("clientId")

            # QA4-022: Validate status against BriefStatus enum
            valid_statuses = [s.value for s in BriefStatus]
            if status not in valid_statuses:
                return JsonResponse(
                    {
                        "error": (
                            f"Invalid status. Must be one of:"
                            f" {', '.join(valid_statuses)}"
                        )
                    },
                    status=400,
                )

            # QA3-015: For CLIENT users, force client_id from authenticated user
            user_group = request.user_data.get("group")
            if user_group == "CLIENT":
                client_id = request.user_data.get("client_id")

            brief = Brief.objects.create(
                status=status,
                details=details,
                client_id=client_id if client_id else None,
            )

            return JsonResponse(serialize_brief(brief), status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error creating brief")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET", "PUT", "PATCH", "DELETE"])
@require_groups("VENDOR", "CLIENT", "SYSTEM")
def brief_detail(request, brief_id):  # noqa: C901, PLR0912
    """Get, update, or delete a specific brief."""
    try:
        brief = Brief.objects.get(id=brief_id, deleted_at__isnull=True)
    except Brief.DoesNotExist:
        return JsonResponse({"error": "Brief not found"}, status=404)

    # QA4-017: Role-aware ownership check
    user_group = request.user_data.get("group")
    if user_group == "CLIENT":
        user_client_id = request.user_data.get("client_id")
        if not user_client_id or str(brief.client_id) != user_client_id:
            return JsonResponse({"error": "Access denied"}, status=403)
    elif user_group == "VENDOR":
        user_vendor_id = request.user_data.get("vendor_id")
        if (
            not user_vendor_id
            or not brief.projects.filter(vendor_id=user_vendor_id).exists()
        ):
            return JsonResponse({"error": "Access denied"}, status=403)
    elif user_group != "SYSTEM":
        return JsonResponse({"error": "Access denied"}, status=403)

    if request.method == "GET":
        return JsonResponse(serialize_brief(brief))

    if request.method in ["PUT", "PATCH"]:
        try:
            # QA2-020: Reject excessively large request bodies
            if len(request.body) > MAX_REQUEST_BODY_SIZE:
                return JsonResponse({"error": "Request body too large"}, status=400)

            data = json.loads(request.body)

            if "status" in data:
                valid_statuses = [s.value for s in BriefStatus]
                if data["status"] not in valid_statuses:
                    return JsonResponse(
                        {
                            "error": (
                                f"Invalid status. Must be one of:"
                                f" {', '.join(valid_statuses)}"
                            )
                        },
                        status=400,
                    )
                brief.status = data["status"]
            if "details" in data:
                brief.details = data["details"]
            # QA2-009: Do NOT allow changing clientId via PATCH
            # to prevent mass assignment

            brief.save()
            return JsonResponse(serialize_brief(brief))

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error updating brief")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    if request.method == "DELETE":
        brief.deleted_at = datetime.now(UTC)
        brief.save()
        return JsonResponse({"message": "Brief deleted"}, status=200)

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ==================== Offers API ====================


@csrf_exempt
@require_http_methods(["GET", "POST"])
@require_groups("VENDOR", "CLIENT", "SYSTEM")
def offers_list(request):  # noqa: C901, PLR0912, PLR0915
    """List all offers or create a new one."""
    if request.method == "GET":
        user_vendor_id = request.user_data.get("vendor_id")
        user_group = request.user_data.get("group")
        # Get project_id from query params if provided
        project_id = request.GET.get("projectId")
        if project_id:
            # QA2-010: Also verify project ownership for VENDOR users
            filter_kwargs = {
                "project_id": project_id,
                "deleted_at__isnull": True,
            }
            if user_group == "VENDOR" and user_vendor_id:
                filter_kwargs["project__vendor_id"] = user_vendor_id
            offers = Offer.objects.filter(
                **filter_kwargs,
            ).prefetch_related("offer_entries")
        elif user_vendor_id:
            offers = Offer.objects.filter(
                project__vendor_id=user_vendor_id,
                deleted_at__isnull=True,
            ).prefetch_related("offer_entries")
        else:
            offers = Offer.objects.none()

        return JsonResponse([serialize_offer(o) for o in offers], safe=False)

    if request.method == "POST":
        try:
            # QA2-020: Reject excessively large request bodies
            if len(request.body) > MAX_REQUEST_BODY_SIZE:
                return JsonResponse({"error": "Request body too large"}, status=400)

            data = json.loads(request.body)
            project_id = data.get("projectId")
            project_name = data.get("projectName")
            status = data.get("status", "DRAFT")
            details = data.get("details", {})
            description = data.get("description", "")
            deadline = data.get("deadline")
            source = data.get("source", "PLATFORM")
            is_locked = data.get("isLocked", False)
            cost = data.get("cost")
            profit = data.get("profit")

            if not project_id or not project_name or not deadline:
                return JsonResponse(
                    {"error": "projectId, projectName, and deadline are required"},
                    status=400,
                )

            # Validate UUID
            try:
                _validate_uuid(project_id, "projectId")
            except ValueError as e:
                return JsonResponse({"error": str(e)}, status=400)

            # Validate status is a valid OfferStatus value
            valid_statuses = [s.value for s in OfferStatus]
            if status not in valid_statuses:
                return JsonResponse(
                    {
                        "error": (
                            f"Invalid status. Must be one of:"
                            f" {', '.join(valid_statuses)}"
                        )
                    },
                    status=400,
                )

            # Verify project exists
            try:
                project = Project.objects.get(id=project_id)
            except Project.DoesNotExist:
                return JsonResponse({"error": "Project not found"}, status=404)

            # QA2-002: Verify requesting vendor owns the project
            user_vendor_id = request.user_data.get("vendor_id")
            if not user_vendor_id or str(project.vendor_id) != user_vendor_id:
                return JsonResponse({"error": "Access denied"}, status=403)

            # Parse deadline
            try:
                deadline_dt = datetime.fromisoformat(deadline.replace("Z", "+00:00"))
            except (ValueError, AttributeError):
                return JsonResponse({"error": "Invalid deadline format"}, status=400)

            create_kwargs: dict = {
                "project": project,
                "project_name": project_name,
                "description": description,
                "status": status,
                "details": details,
                "deadline": deadline_dt,
                "source": source,
                "is_locked": is_locked,
                "cost": cost,
                "profit": profit,
            }

            if "bidDate" in data:
                if data["bidDate"] is not None:
                    try:
                        create_kwargs["bid_date"] = datetime.strptime(  # noqa: DTZ007
                            data["bidDate"], "%Y-%m-%d"
                        ).date()
                    except (ValueError, AttributeError):
                        return JsonResponse(
                            {"error": "Invalid bidDate format"}, status=400
                        )
            if "revision" in data:
                create_kwargs["revision"] = data["revision"]
            if "term" in data:
                create_kwargs["term"] = data["term"]
            if "territory" in data:
                create_kwargs["territory"] = data["territory"]
            if "mediaPlacements" in data:
                create_kwargs["media_placements"] = data["mediaPlacements"]
            if "coverPageNotes" in data:
                create_kwargs["cover_page_notes"] = data["coverPageNotes"]

            vendor_settings = VendorSettings.objects.filter(
                vendor=project.vendor
            ).first()
            if vendor_settings:
                create_kwargs["fringes_percent"] = vendor_settings.fringes_percent
                create_kwargs["handling_percent"] = vendor_settings.handling_percent
                create_kwargs["markup_percent"] = vendor_settings.markup_percent
                create_kwargs["production_insurance_percent"] = (
                    vendor_settings.production_insurance_percent
                )
                create_kwargs["production_fee_percent"] = (
                    vendor_settings.production_fee_percent
                )
                create_kwargs["post_markup_percent"] = (
                    vendor_settings.post_markup_percent
                )
                create_kwargs["post_insurance_percent"] = (
                    vendor_settings.post_insurance_percent
                )
                create_kwargs["post_tax_percent"] = vendor_settings.post_tax_percent

            offer = Offer.objects.create(**create_kwargs)

            if details:
                try:
                    parse_offer_details_to_entries(offer, details)
                except Exception:
                    logger.exception(
                        "Error parsing offer details to entries for offer %s", offer.id
                    )

                recalculate_offer_totals(offer)

            if "deliverables" in data:
                _sync_offer_deliverables(offer, data["deliverables"])
            if "scheduleEntries" in data:
                _sync_offer_schedule_entries(offer, data["scheduleEntries"])

            return JsonResponse(serialize_offer(offer), status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error creating offer")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET", "PUT", "PATCH", "DELETE"])
@require_groups("VENDOR", "CLIENT", "SYSTEM")
def offer_detail(request, offer_id):  # noqa: C901, PLR0912, PLR0915
    """Get, update, or delete a specific offer."""
    try:
        offer = Offer.objects.select_related("project").get(
            id=offer_id, deleted_at__isnull=True
        )
    except Offer.DoesNotExist:
        return JsonResponse({"error": "Offer not found"}, status=404)

    # Verify ownership: offer's project vendor_id must match requesting user's vendor_id
    user_vendor_id = request.user_data.get("vendor_id")
    if (
        not user_vendor_id
        or not offer.project
        or str(offer.project.vendor_id) != user_vendor_id
    ):
        return JsonResponse({"error": "Access denied"}, status=403)

    if request.method == "GET":
        return JsonResponse(serialize_offer(offer))

    if request.method in ["PUT", "PATCH"]:
        try:
            body_len = len(request.body)
            if body_len > MAX_REQUEST_BODY_SIZE:
                logger.warning(
                    "PATCH offer %s: body too large (%d bytes)", offer_id, body_len
                )
                return JsonResponse({"error": "Request body too large"}, status=400)

            data = json.loads(request.body)

            if "projectName" in data:
                offer.project_name = data["projectName"]
            if "description" in data:
                offer.description = data["description"]
            if "status" in data and data["status"] is not None:
                valid_statuses = [s.value for s in OfferStatus]
                if data["status"] not in valid_statuses:
                    return JsonResponse(
                        {
                            "error": (
                                f"Invalid status. Must be one of:"
                                f" {', '.join(valid_statuses)}"
                            )
                        },
                        status=400,
                    )
                offer.status = data["status"]
            if "deadline" in data:
                if data["deadline"] is None:
                    offer.deadline = None
                else:
                    try:
                        offer.deadline = datetime.fromisoformat(
                            data["deadline"].replace("Z", "+00:00"),
                        )
                    except (ValueError, AttributeError):
                        return JsonResponse(
                            {"error": "Invalid deadline format"},
                            status=400,
                        )
            if "source" in data and data["source"] is not None:
                valid_sources = [s.value for s in OfferSource]
                if data["source"] not in valid_sources:
                    return JsonResponse(
                        {
                            "error": (
                                "Invalid source. Must be one of: "
                                f"{', '.join(valid_sources)}"
                            )
                        },
                        status=400,
                    )
                offer.source = data["source"]
            if "isLocked" in data:
                offer.is_locked = data["isLocked"]
            if "cost" in data:
                offer.cost = data["cost"]
            if "profit" in data:
                offer.profit = data["profit"]
            if "bidDate" in data:
                if data["bidDate"] is None:
                    offer.bid_date = None
                else:
                    try:
                        offer.bid_date = datetime.strptime(  # noqa: DTZ007
                            data["bidDate"], "%Y-%m-%d"
                        ).date()
                    except (ValueError, AttributeError):
                        return JsonResponse(
                            {"error": "Invalid bidDate format"}, status=400
                        )
            if "revision" in data:
                offer.revision = data["revision"]
            if "term" in data:
                offer.term = data["term"]
            if "territory" in data:
                offer.territory = data["territory"]
            if "mediaPlacements" in data:
                offer.media_placements = data["mediaPlacements"]
            if "coverPageNotes" in data:
                offer.cover_page_notes = data["coverPageNotes"]
            if "assumptionsExclusions" in data:
                offer.assumptions_exclusions = data["assumptionsExclusions"]

            percent_fields_map = {
                "fringesPercent": "fringes_percent",
                "handlingPercent": "handling_percent",
                "markupPercent": "markup_percent",
                "productionInsurancePercent": "production_insurance_percent",
                "productionFeePercent": "production_fee_percent",
                "postMarkupPercent": "post_markup_percent",
                "postInsurancePercent": "post_insurance_percent",
                "postTaxPercent": "post_tax_percent",
            }
            for json_key, model_field in percent_fields_map.items():
                if json_key in data:
                    setattr(offer, model_field, Decimal(str(data[json_key])))

            if "deliverables" in data:
                _sync_offer_deliverables(offer, data["deliverables"])
            if "scheduleEntries" in data:
                _sync_offer_schedule_entries(offer, data["scheduleEntries"])

            if "details" in data:
                offer.details = data["details"]
                try:
                    parse_offer_details_to_entries(offer, data["details"])
                except Exception:
                    logger.exception(
                        "Error parsing offer details to entries for offer %s", offer.id
                    )
                    offer.save()

                recalculate_offer_totals(offer)
            else:
                offer.save()

            return JsonResponse(serialize_offer(offer))

        except json.JSONDecodeError:
            logger.exception(
                "PATCH offer %s: invalid JSON body (%d bytes)",
                offer_id,
                len(request.body),
            )
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error updating offer %s", offer_id)
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    if request.method == "DELETE":
        offer.deleted_at = datetime.now(UTC)
        offer.save()
        return JsonResponse({"message": "Offer deleted"}, status=200)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET"])
@require_groups("VENDOR", "CLIENT", "SYSTEM")
def offers_by_project(request, project_id):
    """Get all offers for a specific project."""
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return JsonResponse({"error": "Project not found"}, status=404)

    # QA3-002: Verify ownership
    user_group = request.user_data.get("group")
    user_vendor_id = request.user_data.get("vendor_id")
    user_client_id = request.user_data.get("client_id")

    if user_group == "VENDOR":
        if not user_vendor_id or str(project.vendor_id) != user_vendor_id:
            return JsonResponse({"error": "Access denied"}, status=403)
    elif user_group == "CLIENT":
        # CLIENT can only see offers for projects linked to their briefs
        has_access = (
            Brief.objects.filter(
                client_id=user_client_id,
                projects=project,
                deleted_at__isnull=True,
            ).exists()
            if user_client_id
            else False
        )
        if not has_access:
            return JsonResponse({"error": "Access denied"}, status=403)
    elif user_group != "SYSTEM":
        return JsonResponse({"error": "Access denied"}, status=403)

    offers = Offer.objects.filter(
        project=project, deleted_at__isnull=True
    ).prefetch_related("offer_entries")
    return JsonResponse([serialize_offer(o) for o in offers], safe=False)


# ==================== Shares API ====================


@csrf_exempt
@require_http_methods(["GET", "POST"])
@require_groups("VENDOR", "SYSTEM")
def shares_create(request):  # noqa: C901, PLR0912
    """Create a share link for an offer, or find existing share by offerId.

    GET /api/v1/shares?offerId=uuid — find existing share for an offer
    POST /api/v1/shares — create a new share
    Body: {"offerId": "uuid"}
    Returns the share with token.
    Auto-publishes the offer if it is in DRAFT status.
    Reuses existing active share if one already exists.
    """
    if request.method == "GET":
        offer_id = request.GET.get("offerId")
        if not offer_id:
            return JsonResponse(
                {"error": "offerId query parameter is required"}, status=400
            )
        try:
            _validate_uuid(offer_id, "offerId")
        except ValueError as e:
            return JsonResponse({"error": str(e)}, status=400)
        # Verify the offer belongs to this vendor
        user_vendor_id = request.user_data.get("vendor_id")
        share = (
            Share.objects.filter(
                offer_id=offer_id,
                offer__project__vendor_id=user_vendor_id,
            )
            .select_related("offer", "offer__project")
            .first()
        )
        if not share:
            return JsonResponse({"error": "Share not found"}, status=404)
        return JsonResponse(serialize_share(share))

    try:
        data = json.loads(request.body)
        offer_id = data.get("offerId")

        if not offer_id:
            return JsonResponse({"error": "offerId is required"}, status=400)

        try:
            _validate_uuid(offer_id, "offerId")
        except ValueError as e:
            return JsonResponse({"error": str(e)}, status=400)

        try:
            offer = Offer.objects.select_related("project", "project__vendor").get(
                id=offer_id,
                deleted_at__isnull=True,
            )
        except Offer.DoesNotExist:
            return JsonResponse({"error": "Offer not found"}, status=404)

        # Verify the requesting user owns the offer (via vendor)
        user_vendor_id = request.user_data.get("vendor_id")
        if (
            not user_vendor_id
            or not offer.project
            or str(offer.project.vendor_id) != user_vendor_id
        ):
            return JsonResponse({"error": "Access denied"}, status=403)

        # Auto-publish offer if it's still DRAFT
        if offer.status == OfferStatus.DRAFT:
            offer.status = OfferStatus.PUBLISHED
            offer.save(update_fields=["status", "updated_at"])

        # Check for existing active share — reuse it
        existing_share = Share.objects.filter(offer=offer, is_active=True).first()
        if existing_share:
            return JsonResponse(serialize_share(existing_share), status=200)

        # Get creating user
        user_id = request.user_data.get("id")
        user = None
        if user_id:
            with contextlib.suppress(User.DoesNotExist):
                user = User.objects.get(id=user_id)

        # Create new share
        share = Share.objects.create(
            offer=offer,
            created_by=user,
        )

        return JsonResponse(serialize_share(share), status=201)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    except Exception:
        logger.exception("Error creating share")
        return JsonResponse({"error": "An internal error occurred"}, status=500)


@csrf_exempt
@require_http_methods(["GET"])
@public_endpoint
def share_get_public(request, token):
    """Get offer data by share token. NO AUTH REQUIRED (public endpoint).

    Returns full offer details via reconstruct_details_from_entries() plus vendor info.
    """
    try:
        share = Share.objects.select_related(
            "offer",
            "offer__project",
            "offer__project__vendor",
        ).get(token=token)
    except Share.DoesNotExist:
        return JsonResponse({"error": "Share not found"}, status=404)

    if not share.is_active:
        return JsonResponse({"error": "Share link is no longer active"}, status=410)

    # Block access to archived projects
    if (
        share.offer
        and share.offer.project
        and share.offer.project.deleted_at is not None
    ):
        return JsonResponse({"error": "Project is archived"}, status=410)

    # QA2-019: Don't serve draft offers through share links
    if share.offer and share.offer.status == OfferStatus.DRAFT:
        return JsonResponse({"error": "Offer is not available"}, status=404)

    return JsonResponse(serialize_share_public(share))


@csrf_exempt
@require_http_methods(["PATCH", "DELETE"])
@require_groups("VENDOR", "SYSTEM")
def share_manage(request, token):
    """Manage share: toggle active/inactive (PATCH) or deactivate (DELETE).

    Requires vendor auth (must be offer owner).
    """
    try:
        share = Share.objects.select_related(
            "offer",
            "offer__project",
        ).get(token=token)
    except Share.DoesNotExist:
        return JsonResponse({"error": "Share not found"}, status=404)

    # QA2-003: Reject share management when offer has no project
    if share.offer.project is None:
        return JsonResponse(
            {"error": "Cannot manage share for offer without a project"},
            status=400,
        )

    # Verify ownership
    user_vendor_id = request.user_data.get("vendor_id")
    if not user_vendor_id or str(share.offer.project.vendor_id) != user_vendor_id:
        return JsonResponse({"error": "Access denied"}, status=403)

    # PATCH — toggle active/inactive
    if request.method == "PATCH":
        try:
            data = json.loads(request.body)
            if "isActive" in data:
                share.is_active = bool(data["isActive"])
            else:
                share.is_active = not share.is_active
            share.save(update_fields=["is_active", "updated_at"])
            return JsonResponse(serialize_share(share))
        except json.JSONDecodeError:
            share.is_active = not share.is_active
            share.save(update_fields=["is_active", "updated_at"])
            return JsonResponse(serialize_share(share))
        except Exception:
            logger.exception("Error toggling share")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    # DELETE — deactivate
    if request.method == "DELETE":
        share.is_active = False
        share.save(update_fields=["is_active", "updated_at"])
        return JsonResponse({"message": "Share deactivated"})

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ==================== Share Link to Brief ====================


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("CLIENT", "SYSTEM")
def share_link_to_brief(request, token):
    """Link a shared offer to a client's brief.

    Body: {"briefId": "uuid"}
    Creates BriefOffer association.
    """
    try:
        share = Share.objects.select_related("offer").get(token=token)
    except Share.DoesNotExist:
        return JsonResponse({"error": "Share not found"}, status=404)

    if not share.is_active:
        return JsonResponse({"error": "Share link is no longer active"}, status=410)

    try:
        data = json.loads(request.body)
        brief_id = data.get("briefId")

        if not brief_id:
            return JsonResponse({"error": "briefId is required"}, status=400)

        try:
            brief = Brief.objects.get(id=brief_id, deleted_at__isnull=True)
        except Brief.DoesNotExist:
            return JsonResponse({"error": "Brief not found"}, status=404)

        # QA3-014: Positive assertion ownership check (no NULL bypass)
        user_client_id = request.user_data.get("client_id")
        if (
            not user_client_id
            or not brief.client_id
            or str(brief.client_id) != user_client_id
        ):
            return JsonResponse({"error": "Access denied"}, status=403)

        # Get linking user
        user_id = request.user_data.get("id")
        user = None
        if user_id:
            with contextlib.suppress(User.DoesNotExist):
                user = User.objects.get(id=user_id)

        # Create BriefOffer association (or return existing)
        try:
            with transaction.atomic():
                brief_offer = BriefOffer.objects.create(
                    brief=brief,
                    offer=share.offer,
                    linked_by=user,
                )
            return JsonResponse(serialize_brief_offer(brief_offer), status=201)
        except IntegrityError:
            # Already linked
            brief_offer = BriefOffer.objects.get(brief=brief, offer=share.offer)
            return JsonResponse(serialize_brief_offer(brief_offer), status=200)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    except Exception:
        logger.exception("Error linking share to brief")
        return JsonResponse({"error": "An internal error occurred"}, status=500)


# ==================== Offer Status ====================


@csrf_exempt
@require_http_methods(["PATCH"])
@require_groups("VENDOR", "SYSTEM")
def offer_status_update(request, offer_id):
    """Change offer status. Only the offer owner can change status.

    Body: {"status": "PUBLISHED"}
    Valid statuses: DRAFT, PUBLISHED, ARCHIVED.
    """
    try:
        offer = Offer.objects.select_related("project").get(
            id=offer_id,
            deleted_at__isnull=True,
        )
    except Offer.DoesNotExist:
        return JsonResponse({"error": "Offer not found"}, status=404)

    # QA3-004: Positive assertion ownership check (no NULL bypass)
    user_vendor_id = request.user_data.get("vendor_id")
    if (
        not user_vendor_id
        or not offer.project
        or str(offer.project.vendor_id) != user_vendor_id
    ):
        return JsonResponse({"error": "Access denied"}, status=403)

    try:
        data = json.loads(request.body)
        new_status = data.get("status")

        if not new_status:
            return JsonResponse({"error": "status is required"}, status=400)

        valid_statuses = [s.value for s in OfferStatus]
        if new_status not in valid_statuses:
            return JsonResponse(
                {
                    "error": (
                        f"Invalid status. Must be one of: {', '.join(valid_statuses)}"
                    )
                },
                status=400,
            )

        offer.status = new_status
        offer.save(update_fields=["status", "updated_at"])
        return JsonResponse(serialize_offer(offer))

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    except Exception:
        logger.exception("Error updating offer status")
        return JsonResponse({"error": "An internal error occurred"}, status=500)


def _sync_offer_deliverables(offer, deliverables_data):
    existing_ids = set()
    for i, d in enumerate(deliverables_data):
        deliverable_id = d.get("id")
        if deliverable_id:
            try:
                deliverable = OfferDeliverable.objects.get(
                    id=deliverable_id, offer=offer, deleted_at__isnull=True
                )
                deliverable.quantity = d.get("quantity", deliverable.quantity)
                deliverable.duration = d.get("duration", deliverable.duration)
                deliverable.duration_unit = d.get(
                    "durationUnit", deliverable.duration_unit
                )
                deliverable.notes = d.get("notes", deliverable.notes)
                deliverable.sort_order = i
                deliverable.save()
                existing_ids.add(deliverable.id)
            except OfferDeliverable.DoesNotExist:
                deliverable_id = None
        if not deliverable_id:
            new_d = OfferDeliverable.objects.create(
                offer=offer,
                quantity=d.get("quantity", 1),
                duration=d.get("duration", ""),
                duration_unit=d.get("durationUnit", "Sec"),
                notes=d.get("notes", ""),
                sort_order=i,
            )
            existing_ids.add(new_d.id)

    OfferDeliverable.objects.filter(offer=offer, deleted_at__isnull=True).exclude(
        id__in=existing_ids
    ).update(deleted_at=tz.now())


def _sync_offer_schedule_entries(offer, entries_data):
    existing_ids = set()
    for i, e in enumerate(entries_data):
        entry_id = e.get("id")
        if entry_id:
            try:
                entry = OfferScheduleEntry.objects.get(
                    id=entry_id, offer=offer, deleted_at__isnull=True
                )
                entry.phase_type = e.get("phaseType", entry.phase_type)
                entry.days = e.get("days", entry.days)
                entry.hours_per_day = e.get("hoursPerDay", entry.hours_per_day)
                entry.notes = e.get("notes", entry.notes)
                entry.sort_order = i
                entry.save()
                existing_ids.add(entry.id)
            except OfferScheduleEntry.DoesNotExist:
                entry_id = None
        if not entry_id:
            new_e = OfferScheduleEntry.objects.create(
                offer=offer,
                phase_type=e.get("phaseType", "Prep"),
                days=e.get("days", 1),
                hours_per_day=e.get("hoursPerDay", 12),
                notes=e.get("notes", ""),
                sort_order=i,
            )
            existing_ids.add(new_e.id)

    OfferScheduleEntry.objects.filter(offer=offer, deleted_at__isnull=True).exclude(
        id__in=existing_ids
    ).update(deleted_at=tz.now())


# ==================== Copy Offer ====================


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("VENDOR", "SYSTEM")
def offer_copy(request, offer_id):
    """Create a deep copy of an offer (including all OfferEntry records).

    New offer gets status=DRAFT, appends "(Copy)" to project_name.
    """
    try:
        offer = Offer.objects.select_related("project").get(
            id=offer_id,
            deleted_at__isnull=True,
        )
    except Offer.DoesNotExist:
        return JsonResponse({"error": "Offer not found"}, status=404)

    # QA3-004: Positive assertion ownership check (no NULL bypass)
    user_vendor_id = request.user_data.get("vendor_id")
    if (
        not user_vendor_id
        or not offer.project
        or str(offer.project.vendor_id) != user_vendor_id
    ):
        return JsonResponse({"error": "Access denied"}, status=403)

    try:
        # QA2-021: Wrap entire copy logic in a transaction
        with transaction.atomic():
            # Get all offer entries before creating the copy
            original_entries = list(
                OfferEntry.objects.filter(
                    offer=offer, deleted_at__isnull=True
                ).order_by("sort_order")
            )

            new_offer = Offer.objects.create(
                project_name=f"{offer.project_name} (Copy)",
                project=offer.project,
                parent_offer=offer,
                description=offer.description,
                status=OfferStatus.DRAFT,
                cost=offer.cost,
                profit=offer.profit,
                details=offer.details,
                metadata=offer.metadata,
                deadline=offer.deadline,
                source=offer.source,
                is_locked=False,
                bid_date=offer.bid_date,
                revision=offer.revision,
                term=offer.term,
                territory=offer.territory,
                media_placements=offer.media_placements,
                cover_page_notes=offer.cover_page_notes,
            )

            for entry in original_entries:
                OfferEntry.objects.create(
                    offer=new_offer,
                    frontend_id=entry.frontend_id,
                    item_name=entry.item_name,
                    entry=entry.entry,
                    category=entry.category,
                    price=entry.price,
                    cost=entry.cost,
                    client_price=entry.client_price,
                    client_cost=entry.client_cost,
                    surcharge=entry.surcharge,
                    tax_rate=entry.tax_rate,
                    tax_price=entry.tax_price,
                    show_tax=entry.show_tax,
                    is_linked_surcharge=entry.is_linked_surcharge,
                    market_range=entry.market_range,
                    item_data=entry.item_data,
                    sort_order=entry.sort_order,
                )

            for d in OfferDeliverable.objects.filter(
                offer=offer, deleted_at__isnull=True
            ).order_by("sort_order"):
                OfferDeliverable.objects.create(
                    offer=new_offer,
                    quantity=d.quantity,
                    duration=d.duration,
                    duration_unit=d.duration_unit,
                    notes=d.notes,
                    sort_order=d.sort_order,
                )

            for s in OfferScheduleEntry.objects.filter(
                offer=offer, deleted_at__isnull=True
            ).order_by("sort_order"):
                OfferScheduleEntry.objects.create(
                    offer=new_offer,
                    phase_type=s.phase_type,
                    days=s.days,
                    hours_per_day=s.hours_per_day,
                    notes=s.notes,
                    sort_order=s.sort_order,
                )

            recalculate_offer_totals(new_offer)

        return JsonResponse(serialize_offer(new_offer), status=201)

    except Exception:
        logger.exception("Error copying offer")
        return JsonResponse({"error": "An internal error occurred"}, status=500)


# ==================== Templates API ====================


@csrf_exempt
@require_http_methods(["GET", "POST"])
@require_groups("VENDOR", "SYSTEM")
def templates_list(request):
    """List vendor's templates or create a new one from an offer."""
    vendor_id = request.user_data.get("vendor_id")
    if not vendor_id:
        return JsonResponse({"error": "Vendor context required"}, status=400)

    if request.method == "GET":
        templates = Template.objects.filter(
            vendor_id=vendor_id,
            deleted_at__isnull=True,
        )
        return JsonResponse(
            [serialize_template(t) for t in templates],
            safe=False,
        )

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            offer_id = data.get("offerId")
            name = data.get("name")

            if not offer_id or not name:
                return JsonResponse(
                    {"error": "offerId and name are required"},
                    status=400,
                )

            # Verify the offer exists and belongs to this vendor
            try:
                offer = Offer.objects.select_related("project").get(
                    id=offer_id,
                    deleted_at__isnull=True,
                )
            except Offer.DoesNotExist:
                return JsonResponse({"error": "Offer not found"}, status=404)

            # QA4-030: offer must have a project owned by this vendor
            if not offer.project or str(offer.project.vendor_id) != vendor_id:
                return JsonResponse({"error": "Access denied"}, status=403)

            # Snapshot the full offer details and OfferEntry data
            entries_qs = OfferEntry.objects.filter(
                offer=offer,
                deleted_at__isnull=True,
            ).order_by("sort_order")
            entries_count = entries_qs.count()

            if entries_count > 0:
                # Reconstruct details from OfferEntry records
                details = reconstruct_details_from_entries(offer)

                # Snapshot OfferEntry data into the template for completeness
                entries_snapshot = [
                    {
                        "frontendId": entry.frontend_id,
                        "itemName": entry.item_name,
                        "entryId": str(entry.entry_id) if entry.entry_id else None,
                        "categoryId": str(entry.category_id)
                        if entry.category_id
                        else None,
                        "price": str(entry.price) if entry.price is not None else None,
                        "cost": str(entry.cost) if entry.cost is not None else None,
                        "clientPrice": str(entry.client_price)
                        if entry.client_price is not None
                        else None,
                        "clientCost": str(entry.client_cost)
                        if entry.client_cost is not None
                        else None,
                        "surcharge": str(entry.surcharge)
                        if entry.surcharge is not None
                        else None,
                        "taxRate": str(entry.tax_rate),
                        "taxPrice": str(entry.tax_price)
                        if entry.tax_price is not None
                        else None,
                        "showTax": entry.show_tax,
                        "isLinkedSurcharge": entry.is_linked_surcharge,
                        "marketRange": entry.market_range,
                        "itemData": entry.item_data,
                        "sortOrder": entry.sort_order,
                    }
                    for entry in entries_qs
                ]
            else:
                # Fallback: use raw offer.details if no OfferEntry records exist
                details = offer.details if offer.details else {}
                entries_snapshot = []

            metadata = {
                "sourceOfferName": offer.project_name,
                "sourceProjectId": str(offer.project_id) if offer.project_id else None,
                "entriesSnapshot": entries_snapshot,
                "offerMetadata": offer.metadata,
            }

            template = Template.objects.create(
                name=name,
                vendor_id=vendor_id,
                source_offer=offer,
                details=details,
                description=offer.description,
                metadata=metadata,
            )

            return JsonResponse(serialize_template(template), status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error creating template")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "DELETE"])
@require_groups("VENDOR", "SYSTEM")
def template_detail(request, template_id):  # noqa: C901
    """Get, update, or delete a specific template."""
    vendor_id = request.user_data.get("vendor_id")
    if not vendor_id:
        return JsonResponse({"error": "Vendor context required"}, status=400)

    try:
        template = Template.objects.get(
            id=template_id,
            vendor_id=vendor_id,
            deleted_at__isnull=True,
        )
    except Template.DoesNotExist:
        return JsonResponse({"error": "Template not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(serialize_template(template))

    if request.method == "PATCH":
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        # Update only the fields that are provided
        if "name" in data:
            template.name = data["name"]
        if "description" in data:
            template.description = data["description"]
        if "metadata" in data:
            template.metadata = data["metadata"]

        if "details" in data:
            template.details = data["details"]

            # Recalculate entriesSnapshot from the new details
            offers_list = data["details"].get("offers", [])
            entries_snapshot = []
            for idx, item in enumerate(offers_list):
                # Extract known structured fields; everything else goes into itemData
                known_keys = {
                    "id",
                    "item",
                    "entryId",
                    "categoryId",
                    "price",
                    "cost",
                    "clientPrice",
                    "clientCost",
                    "surcharge",
                    "taxRate",
                    "taxPrice",
                    "showTax",
                    "isLinkedSurcharge",
                    "marketRange",
                }
                item_data = {k: v for k, v in item.items() if k not in known_keys}

                entries_snapshot.append(
                    {
                        "frontendId": item.get("id"),
                        "itemName": item.get("item", ""),
                        "entryId": item.get("entryId"),
                        "categoryId": item.get("categoryId"),
                        "price": str(item["price"])
                        if item.get("price") is not None
                        else None,
                        "cost": str(item["cost"])
                        if item.get("cost") is not None
                        else None,
                        "clientPrice": str(item["clientPrice"])
                        if item.get("clientPrice") is not None
                        else None,
                        "clientCost": str(item["clientCost"])
                        if item.get("clientCost") is not None
                        else None,
                        "surcharge": str(item["surcharge"])
                        if item.get("surcharge") is not None
                        else None,
                        "taxRate": str(item.get("taxRate", 0)),
                        "taxPrice": str(item["taxPrice"])
                        if item.get("taxPrice") is not None
                        else None,
                        "showTax": item.get("showTax", False),
                        "isLinkedSurcharge": item.get("isLinkedSurcharge", False),
                        "marketRange": item.get("marketRange"),
                        "itemData": item_data if item_data else None,
                        "sortOrder": idx,
                    }
                )

            # Store entriesSnapshot inside metadata
            if not isinstance(template.metadata, dict):
                template.metadata = {}
            template.metadata["entriesSnapshot"] = entries_snapshot

        template.save()
        return JsonResponse(serialize_template(template))

    if request.method == "DELETE":
        template.deleted_at = datetime.now(UTC)
        template.save()
        return JsonResponse({"message": "Template deleted"}, status=200)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("VENDOR", "SYSTEM")
def template_apply(request, template_id):
    """Create a new offer from a template.

    Body: {"projectId": "uuid"}
    Creates Offer + OfferEntry records from template snapshot.
    """
    vendor_id = request.user_data.get("vendor_id")
    if not vendor_id:
        return JsonResponse({"error": "Vendor context required"}, status=400)

    try:
        template = Template.objects.get(
            id=template_id,
            vendor_id=vendor_id,
            deleted_at__isnull=True,
        )
    except Template.DoesNotExist:
        return JsonResponse({"error": "Template not found"}, status=404)

    try:
        data = json.loads(request.body)
        project_id = data.get("projectId")

        if not project_id:
            return JsonResponse({"error": "projectId is required"}, status=400)

        try:
            project = Project.objects.get(
                id=project_id,
                vendor_id=vendor_id,
                deleted_at__isnull=True,
            )
        except Project.DoesNotExist:
            return JsonResponse({"error": "Project not found"}, status=404)

        with transaction.atomic():
            # Create the offer from template
            offer = Offer.objects.create(
                project=project,
                project_name=project.name,
                description=template.description,
                status=OfferStatus.DRAFT,
                details=template.details,
                metadata=template.metadata.get("offerMetadata", {}),
                deadline=None,
                source=OfferSource.PLATFORM,
                is_locked=False,
            )

            # Create OfferEntry records from entries snapshot
            entries_snapshot = template.metadata.get("entriesSnapshot", [])

            entry_ids = {e["entryId"] for e in entries_snapshot if e.get("entryId")}
            category_ids = {
                e["categoryId"] for e in entries_snapshot if e.get("categoryId")
            }
            entries_by_id = (
                {str(e.id): e for e in Entry.objects.filter(id__in=entry_ids)}
                if entry_ids
                else {}
            )
            categories_by_id = (
                {str(c.id): c for c in Category.objects.filter(id__in=category_ids)}
                if category_ids
                else {}
            )

            for entry_data in entries_snapshot:
                entry_ref = entries_by_id.get(entry_data.get("entryId"))
                category_ref = categories_by_id.get(entry_data.get("categoryId"))

                OfferEntry.objects.create(
                    offer=offer,
                    frontend_id=entry_data.get("frontendId", ""),
                    item_name=entry_data.get("itemName", ""),
                    entry=entry_ref,
                    category=category_ref,
                    price=entry_data.get("price"),
                    cost=entry_data.get("cost"),
                    client_price=entry_data.get("clientPrice"),
                    client_cost=entry_data.get("clientCost"),
                    surcharge=entry_data.get("surcharge"),
                    tax_rate=entry_data.get("taxRate", 0),
                    tax_price=entry_data.get("taxPrice"),
                    show_tax=entry_data.get("showTax", False),
                    is_linked_surcharge=entry_data.get("isLinkedSurcharge", True),
                    market_range=entry_data.get("marketRange", ""),
                    item_data=entry_data.get("itemData", {}),
                    sort_order=entry_data.get("sortOrder", 0),
                )

            recalculate_offer_totals(offer)

        return JsonResponse(serialize_offer(offer), status=201)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    except Exception:
        logger.exception("Error applying template")
        return JsonResponse({"error": "An internal error occurred"}, status=500)


# ==================== Rate Cards API ====================


@csrf_exempt
@require_http_methods(["GET", "POST"])
@require_groups("VENDOR", "SYSTEM")
def rate_cards_list(request):
    """List vendor's rate cards or create a new one."""
    vendor_id = request.user_data.get("vendor_id")
    if not vendor_id:
        return JsonResponse({"error": "Vendor context required"}, status=400)

    if request.method == "GET":
        rate_cards = RateCard.objects.filter(
            vendor_id=vendor_id,
            deleted_at__isnull=True,
        )
        return JsonResponse(
            [serialize_rate_card(rc) for rc in rate_cards],
            safe=False,
        )

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            name = data.get("name")
            items = data.get("items", [])

            if not name:
                return JsonResponse({"error": "name is required"}, status=400)

            with transaction.atomic():
                rate_card = RateCard.objects.create(
                    vendor_id=vendor_id,
                    name=name,
                )

                for item_data in items:
                    item_name = item_data.get("itemName", "")
                    price = item_data.get("price", 0)
                    entry_id = item_data.get("entryId")
                    unit_id = item_data.get("unitId")
                    unit_label = item_data.get("unitLabel", "")

                    entry_ref = None
                    if entry_id:
                        with contextlib.suppress(Entry.DoesNotExist):
                            entry_ref = Entry.objects.get(id=entry_id)

                    unit_ref = None
                    if unit_id:
                        with contextlib.suppress(Unit.DoesNotExist):
                            unit_ref = Unit.objects.get(id=unit_id)

                    RateCardItem.objects.create(
                        rate_card=rate_card,
                        entry=entry_ref,
                        item_name=item_name,
                        price=price,
                        unit=unit_ref,
                        unit_label=unit_label,
                    )

            return JsonResponse(serialize_rate_card(rate_card), status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error creating rate card")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "DELETE"])
@require_groups("VENDOR", "SYSTEM")
def rate_card_detail(request, rate_card_id):  # noqa: C901
    """Get, update, or delete a specific rate card."""
    vendor_id = request.user_data.get("vendor_id")
    if not vendor_id:
        return JsonResponse({"error": "Vendor context required"}, status=400)

    try:
        rate_card = RateCard.objects.get(
            id=rate_card_id,
            vendor_id=vendor_id,
            deleted_at__isnull=True,
        )
    except RateCard.DoesNotExist:
        return JsonResponse({"error": "Rate card not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(serialize_rate_card(rate_card))

    if request.method == "PATCH":
        try:
            data = json.loads(request.body)

            if "name" in data:
                rate_card.name = data["name"]
                rate_card.save()

            # Replace items if provided
            if "items" in data:
                with transaction.atomic():
                    # Hard delete existing items and recreate
                    rate_card.items.all().delete()
                    for item_data in data["items"]:
                        item_name = item_data.get("itemName", "")
                        price = item_data.get("price", 0)
                        entry_id = item_data.get("entryId")
                        unit_id = item_data.get("unitId")
                        unit_label = item_data.get("unitLabel", "")

                        entry_ref = None
                        if entry_id:
                            with contextlib.suppress(Entry.DoesNotExist):
                                entry_ref = Entry.objects.get(id=entry_id)

                        unit_ref = None
                        if unit_id:
                            with contextlib.suppress(Unit.DoesNotExist):
                                unit_ref = Unit.objects.get(id=unit_id)

                        RateCardItem.objects.create(
                            rate_card=rate_card,
                            entry=entry_ref,
                            item_name=item_name,
                            price=price,
                            unit=unit_ref,
                            unit_label=unit_label,
                        )

            return JsonResponse(serialize_rate_card(rate_card))

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error updating rate card")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    if request.method == "DELETE":
        rate_card.deleted_at = datetime.now(UTC)
        rate_card.save()
        # Also soft-delete items
        rate_card.items.filter(deleted_at__isnull=True).update(
            deleted_at=datetime.now(UTC)
        )
        return JsonResponse({"message": "Rate card deleted"}, status=200)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET"])
@require_groups("VENDOR", "SYSTEM")
def rate_card_lookup(request):
    """Lookup rate for an entry across vendor's rate cards.

    Query params: entryId=uuid
    Returns matching rate card items for the given entry.
    """
    vendor_id = request.user_data.get("vendor_id")
    if not vendor_id:
        return JsonResponse({"error": "Vendor context required"}, status=400)

    entry_id = request.GET.get("entryId")
    if not entry_id:
        return JsonResponse(
            {"error": "entryId query parameter is required"}, status=400
        )

    items = RateCardItem.objects.filter(
        rate_card__vendor_id=vendor_id,
        rate_card__deleted_at__isnull=True,
        entry_id=entry_id,
        deleted_at__isnull=True,
    ).select_related("rate_card")

    results = []
    for item in items:
        item_data = serialize_rate_card_item(item)
        item_data["rateCardName"] = item.rate_card.name
        results.append(item_data)

    return JsonResponse(results, safe=False)


# ==================== Client Briefs API ====================


@csrf_exempt
@require_http_methods(["GET", "POST"])
@require_groups("CLIENT", "SYSTEM")
def client_briefs_list(request):
    """List client's briefs with offers count, or create a new one."""
    client_id = request.user_data.get("client_id")
    if not client_id:
        return JsonResponse({"error": "Client context required"}, status=400)

    if request.method == "GET":
        briefs = Brief.objects.filter(
            client_id=client_id,
            deleted_at__isnull=True,
        )
        return JsonResponse(
            [serialize_brief_with_offers(b) for b in briefs],
            safe=False,
        )

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            name = data.get("name", "")
            details = data.get("details", {})
            status = data.get("status", "DRAFT")

            # QA4-022: Validate status against BriefStatus enum
            valid_statuses = [s.value for s in BriefStatus]
            if status not in valid_statuses:
                return JsonResponse(
                    {
                        "error": (
                            f"Invalid status. Must be one of:"
                            f" {', '.join(valid_statuses)}"
                        )
                    },
                    status=400,
                )

            # Store name inside details if provided
            if name and isinstance(details, dict):
                details["name"] = name

            brief = Brief.objects.create(
                status=status,
                details=details,
                client_id=client_id,
            )

            return JsonResponse(serialize_brief_with_offers(brief), status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error creating brief")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET", "PATCH", "DELETE"])
@require_groups("CLIENT", "SYSTEM")
def client_brief_detail(request, brief_id):  # noqa: C901, PLR0912
    """Get, update, or delete a specific client brief."""
    client_id = request.user_data.get("client_id")
    if not client_id:
        return JsonResponse({"error": "Client context required"}, status=400)

    try:
        brief = Brief.objects.get(
            id=brief_id,
            client_id=client_id,
            deleted_at__isnull=True,
        )
    except Brief.DoesNotExist:
        return JsonResponse({"error": "Brief not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(serialize_brief_detail(brief))

    if request.method == "PATCH":
        try:
            data = json.loads(request.body)

            if "details" in data:
                brief.details = data["details"]
            if "name" in data:
                if isinstance(brief.details, dict):
                    brief.details["name"] = data["name"]
                else:
                    brief.details = {"name": data["name"]}
            if "status" in data:
                valid_statuses = [s.value for s in BriefStatus]
                if data["status"] not in valid_statuses:
                    return JsonResponse(
                        {
                            "error": (
                                f"Invalid status. Must be one of:"
                                f" {', '.join(valid_statuses)}"
                            )
                        },
                        status=400,
                    )
                brief.status = data["status"]

            brief.save()
            return JsonResponse(serialize_brief_detail(brief))

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception:
            logger.exception("Error updating brief")
            return JsonResponse({"error": "An internal error occurred"}, status=500)

    if request.method == "DELETE":
        brief.deleted_at = datetime.now(UTC)
        brief.save()
        return JsonResponse({"message": "Brief deleted"}, status=200)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["GET"])
@require_groups("CLIENT", "SYSTEM")
def client_brief_offers(request, brief_id):
    """Get all offers linked to a brief (via BriefOffer model)."""
    client_id = request.user_data.get("client_id")
    if not client_id:
        return JsonResponse({"error": "Client context required"}, status=400)

    try:
        brief = Brief.objects.get(
            id=brief_id,
            client_id=client_id,
            deleted_at__isnull=True,
        )
    except Brief.DoesNotExist:
        return JsonResponse({"error": "Brief not found"}, status=404)

    brief_offers = BriefOffer.objects.filter(
        brief=brief,
    ).select_related(
        "offer",
        "offer__project",
        "offer__project__vendor",
    )

    offers = []
    for bo in brief_offers:
        offer = bo.offer
        # QA4-023: Use client serializer to exclude cost/profit
        offer_data = serialize_offer_for_client(offer)
        offer_data["linkedAt"] = bo.created_at.isoformat() if bo.created_at else None
        offer_data["linkedBy"] = str(bo.linked_by_id) if bo.linked_by_id else None
        if offer.project and offer.project.vendor:
            offer_data["vendor"] = {
                "id": str(offer.project.vendor.id),
                "name": offer.project.vendor.name,
            }
        else:
            offer_data["vendor"] = None
        offers.append(offer_data)

    return JsonResponse(offers, safe=False)


# ==================== AI Brief Chat API ====================


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("CLIENT", "SYSTEM")
@ratelimit(key="user_or_ip", rate="20/m", method="POST", block=True)
def client_brief_chat(request):
    """AI-powered chat for brief creation."""
    try:
        data = json.loads(request.body)
        user_message = data.get("message", "")
        history = data.get("history", [])
        brief_id = data.get("brief_id")
        extracted_fields = data.get("extracted_fields", {})

        if not user_message:
            return JsonResponse({"error": "message is required"}, status=400)

        # Get user for storing chat messages
        user_id = request.user_data.get("id")
        user = None
        if user_id:
            with contextlib.suppress(User.DoesNotExist):
                user = User.objects.get(id=user_id)

        # QA3-039: Get brief with ownership check
        client_id = request.user_data.get("client_id")
        brief = None
        if brief_id:
            with contextlib.suppress(Brief.DoesNotExist):
                brief = Brief.objects.get(
                    id=brief_id, client_id=client_id, deleted_at__isnull=True
                )

        # Store user message
        if user:
            ChatMessage.objects.create(
                brief=brief,
                user=user,
                role="user",
                content=user_message,
                metadata={"extracted_fields_before": extracted_fields},
            )

        # Process through LangGraph
        result = process_chat_message(user_message, history, extracted_fields)

        # Store assistant reply
        if user:
            ChatMessage.objects.create(
                brief=brief,
                user=user,
                role="assistant",
                content=result["reply"],
                metadata={
                    "extracted_fields": result.get("extracted_fields", {}),
                    "is_complete": result.get("is_complete", False),
                },
            )

        return JsonResponse(result)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    except Exception:
        logger.exception("Error in brief chat")
        return JsonResponse({"error": "An internal error occurred"}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("CLIENT", "SYSTEM")
@ratelimit(key="user_or_ip", rate="10/m", method="POST", block=True)
def client_brief_chat_analyze(request):
    """Analyze a brief and provide suggestions.

    Body: {"brief_data": {...}}
    Returns: {"suggestions": [...], "summary": "..."}
    """
    try:
        data = json.loads(request.body)
        brief_data = data.get("brief_data", {})

        if not brief_data:
            return JsonResponse({"error": "brief_data is required"}, status=400)

        result = analyze_brief(brief_data)
        return JsonResponse(result)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    except Exception:
        logger.exception("Error analyzing brief")
        return JsonResponse({"error": "An internal error occurred"}, status=500)


# ==================== Comparison API ====================


def _build_comparison_data(brief):  # noqa: C901
    """Build comparison data from all offers linked to a brief.

    Returns structured data with vendors, categories, items, and totals.
    """
    # QA3-036: prefetch offer entries to avoid N+1 queries
    brief_offers = (
        BriefOffer.objects.filter(
            brief=brief,
        )
        .select_related(
            "offer",
            "offer__project",
            "offer__project__vendor",
        )
        .prefetch_related(
            "offer__offer_entries",
            "offer__offer_entries__category",
            "offer__offer_entries__entry",
        )
    )

    if not brief_offers.exists():
        return {
            "brief": serialize_brief(brief),
            "vendors": [],
            "categories": [],
            "grand_totals": [],
        }

    # Collect vendor info and their offer entries
    vendors = []
    vendor_entries = {}  # vendor_id -> list of OfferEntry

    for bo in brief_offers:
        offer = bo.offer
        vendor_id = None
        vendor_name = "Unknown Vendor"

        if offer.project and offer.project.vendor:
            vendor_id = str(offer.project.vendor.id)
            vendor_name = offer.project.vendor.name
        else:
            vendor_id = str(offer.id)  # Use offer ID as fallback
            vendor_name = offer.project_name

        # QA3-036: Use prefetched entries instead of per-offer query
        entries = sorted(
            [e for e in offer.offer_entries.all() if e.deleted_at is None],
            key=lambda e: e.sort_order,
        )

        vendor_total = sum((e.client_price or e.price or Decimal("0")) for e in entries)

        vendors.append(
            {
                "id": vendor_id,
                "name": vendor_name,
                "offerId": str(offer.id),
                "total": float(vendor_total),
            }
        )

        vendor_entries[vendor_id] = list(entries)

    # Build categories with items and per-vendor values
    # Collect all unique categories across all vendors
    all_categories: dict = {}
    category_items: defaultdict[str, dict] = defaultdict(dict)

    for vendor_id, entries in vendor_entries.items():
        for entry in entries:
            cat_id = str(entry.category_id) if entry.category_id else "uncategorized"
            cat_name = entry.category.name if entry.category else "Uncategorized"
            all_categories[cat_id] = cat_name

            item_name = entry.item_name or (
                entry.entry.name if entry.entry else f"Item {entry.sort_order}"
            )
            if item_name not in category_items[cat_id]:
                category_items[cat_id][item_name] = {}

            category_items[cat_id][item_name][vendor_id] = {
                "vendor_id": vendor_id,
                "price": float(entry.client_price or entry.price or 0),
                "cost": float(entry.client_cost or entry.cost or 0),
            }

    # Build structured categories response
    categories = []
    grand_totals = {v["id"]: 0.0 for v in vendors}

    for cat_id, cat_name in all_categories.items():
        items_data = []
        subtotals = {v["id"]: 0.0 for v in vendors}

        for item_name, vendor_values in category_items[cat_id].items():
            values = []
            for v in vendors:
                v_data = vendor_values.get(
                    v["id"], {"vendor_id": v["id"], "price": 0, "cost": 0}
                )
                values.append(v_data)
                subtotals[v["id"]] += v_data.get("price", 0)

            items_data.append(
                {
                    "name": item_name,
                    "values": values,
                }
            )

        # Add subtotals to grand totals
        for v_id, subtotal in subtotals.items():
            grand_totals[v_id] += subtotal

        categories.append(
            {
                "id": cat_id,
                "name": cat_name,
                "items": items_data,
                "subtotals": [
                    {"vendor_id": v["id"], "total": subtotals[v["id"]]} for v in vendors
                ],
            }
        )

    return {
        "brief": serialize_brief(brief),
        "vendors": vendors,
        "categories": categories,
        "grand_totals": [
            {"vendor_id": v["id"], "total": grand_totals[v["id"]]} for v in vendors
        ],
    }


@csrf_exempt
@require_http_methods(["GET"])
@require_groups("CLIENT", "SYSTEM")
def client_brief_comparison(request, brief_id):
    """Get comparison data for all offers linked to a brief.

    Returns aggregated data from all offers grouped by category and item.
    """
    client_id = request.user_data.get("client_id")
    if not client_id:
        return JsonResponse({"error": "Client context required"}, status=400)

    try:
        brief = Brief.objects.get(
            id=brief_id,
            client_id=client_id,
            deleted_at__isnull=True,
        )
    except Brief.DoesNotExist:
        return JsonResponse({"error": "Brief not found"}, status=404)

    comparison_data = _build_comparison_data(brief)
    return JsonResponse(comparison_data)


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("CLIENT", "SYSTEM")
@ratelimit(key="user_or_ip", rate="10/m", method="POST", block=True)
def client_brief_comparison_analyze(request, brief_id):
    """AI analysis of comparison data for a brief's offers.

    Body: {"question": "..."} (optional - for follow-up questions)
    Returns: {"analysis": "...", "highlights": [...]}
    """
    client_id = request.user_data.get("client_id")
    if not client_id:
        return JsonResponse({"error": "Client context required"}, status=400)

    try:
        brief = Brief.objects.get(
            id=brief_id,
            client_id=client_id,
            deleted_at__isnull=True,
        )
    except Brief.DoesNotExist:
        return JsonResponse({"error": "Brief not found"}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = {}

    question = data.get("question")

    # Build comparison data
    comparison_data = _build_comparison_data(brief)

    if not comparison_data.get("vendors"):
        return JsonResponse(
            {
                "analysis": (
                    "No vendor offers are linked to this brief yet."
                    " Link some offers first to get a comparison analysis."
                ),
                "highlights": [],
            }
        )

    # Use AI to analyze
    result = analyze_comparison(
        brief_data=brief.details,
        comparison_data=comparison_data,
        question=question,
    )

    return JsonResponse(result)


# ==================== XLSX Upload API ====================


@csrf_exempt
@require_http_methods(["POST"])
@require_groups("CLIENT", "SYSTEM")
def client_xlsx_upload(request):  # noqa: C901, PLR0912
    """Upload XLSX file, find offer_id cell, return share info.

    Scans all cells in all sheets for a value that looks like a UUID.
    If the UUID matches an existing Offer, returns share information.
    """
    file = request.FILES.get("file")
    if not file:
        return JsonResponse({"error": "No file provided"}, status=400)

    # Validate file size (max 10MB)
    if file.size > 10 * 1024 * 1024:
        return JsonResponse({"error": "File size must not exceed 10MB"}, status=400)

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception:
        logger.exception("Error reading XLSX file")
        return JsonResponse({"error": "Invalid or corrupted XLSX file"}, status=400)

    offer_id = None
    offer = None
    max_sheets = 20
    max_cells_per_sheet = 10000
    try:
        for sheet_idx, sheet in enumerate(wb.sheetnames):
            if sheet_idx >= max_sheets:
                break
            ws = wb[sheet]
            cell_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    cell_count += 1
                    if cell_count > max_cells_per_sheet:
                        break
                    value = str(cell.value or "").strip()
                    # Check if it looks like a UUID
                    try:
                        parsed = uuid_module.UUID(value)
                        # Check if this UUID is an offer
                        found_offer = Offer.objects.filter(
                            id=parsed,
                            deleted_at__isnull=True,
                        ).first()
                        if found_offer:
                            offer_id = str(found_offer.id)
                            offer = found_offer
                            break
                    except (ValueError, AttributeError):
                        continue
                if offer_id or cell_count > max_cells_per_sheet:
                    break
            if offer_id:
                break
    finally:
        wb.close()

    if not offer_id or not offer:
        return JsonResponse(
            {"error": "No valid offer ID found in the file"},
            status=404,
        )

    # QA2-018: Verify the client has access via an existing share
    share = Share.objects.filter(offer_id=offer_id, is_active=True).first()
    if not share:
        return JsonResponse(
            {"error": "No active share found for this offer"},
            status=403,
        )

    return JsonResponse(
        {
            "offer_id": offer_id,
            "offer_name": offer.project_name,
            "share_token": share.token if share else None,
            "has_share": share is not None,
        }
    )


def _build_offer_export_data(offer):
    project = offer.project
    vendor = project.vendor

    vendor_settings = VendorSettings.objects.filter(vendor=vendor).first()
    logo_url = None
    if vendor_settings and vendor_settings.logo:
        logo_url = vendor_settings.logo.url

    collaborators = ProjectCollaborator.objects.filter(
        project=project,
    ).values("id", "name", "email", "role", "user_id")
    collaborators_list = [
        {
            "id": str(x["id"]),
            "userId": str(x["user_id"]) if x["user_id"] else None,
            "name": x["name"],
            "email": x["email"],
            "role": x["role"],
        }
        for x in collaborators
    ]

    client_managers = ClientManager.objects.filter(
        project=project,
    ).values("id", "name", "position")
    client_managers_list = [
        {
            "id": str(x["id"]),
            "name": x["name"],
            "position": x["position"],
        }
        for x in client_managers
    ]

    entries = (
        OfferEntry.objects.filter(
            offer=offer,
            deleted_at__isnull=True,
        )
        .select_related("category", "category__parent_category", "entry")
        .order_by("sort_order")
    )

    categories_map: dict[str, dict] = {}
    for entry in entries:
        category_id = str(entry.category_id) if entry.category_id else "uncategorized"
        if category_id not in categories_map:
            parent = entry.category.parent_category if entry.category else None
            categories_map[category_id] = {
                "id": category_id
                if category_id != "uncategorized"
                else "uncategorized",
                "code": entry.category.code if entry.category else "",
                "name": entry.category.name if entry.category else "Uncategorized",
                "level": entry.category.level if entry.category else 0,
                "parentCategoryId": str(parent.id) if parent else None,
                "parentCategoryName": parent.name if parent else None,
                "parentCategoryCode": parent.code if parent else None,
                "tags": entry.category.tags if entry.category else [],
                "parentTags": parent.tags if parent else [],
                "entries": [],
                "_raw_estimates": [],
            }

        estimate = (
            float(entry.client_cost)
            if entry.client_cost is not None
            else float(entry.cost or 0)
        )
        rate = float(entry.price) if entry.price is not None else 0

        item_data = entry.item_data or {}
        raw_units = item_data.get("units") or []
        units_list = [
            {
                "label": u.get("label", ""),
                "symbol": u.get("label", ""),
                "count": u.get("count", 0),
            }
            for u in raw_units
            if u is not None
        ]

        categories_map[category_id]["entries"].append(
            {
                "id": str(entry.id),
                "entryId": str(entry.entry_id) if entry.entry_id else str(entry.id),
                "code": entry.entry.code if entry.entry else "",
                "name": entry.item_name,
                "rate": rate,
                "units": units_list,
                "overtime": float(entry.overtime) if entry.overtime else 0,
                "estimate": estimate,
            }
        )
        categories_map[category_id]["_raw_estimates"].append(estimate)

    categories_list = []
    for cat in categories_map.values():
        sub_total = sum(cat["_raw_estimates"])
        del cat["_raw_estimates"]
        cat["subTotal"] = round(sub_total, 2)
        cat["fringes"] = None
        cat["sectionTotal"] = round(sub_total, 2)
        categories_list.append(cat)

    share = Share.objects.filter(offer=offer, is_active=True).first()

    deliverables = OfferDeliverable.objects.filter(
        offer=offer,
        deleted_at__isnull=True,
    ).order_by("sort_order")

    schedule_entries = OfferScheduleEntry.objects.filter(
        offer=offer,
        deleted_at__isnull=True,
    ).order_by("sort_order")

    company_name = vendor_settings.company_name if vendor_settings else ""

    return {
        "offer": {
            "id": str(offer.id),
            "uuid": str(offer.id),
            "status": offer.status,
            "cost": float(offer.cost) if offer.cost is not None else None,
            "bidDate": offer.bid_date.isoformat() if offer.bid_date else None,
            "revision": offer.revision,
            "term": offer.term,
            "territory": offer.territory,
            "mediaPlacements": offer.media_placements,
            "coverPageNotes": offer.cover_page_notes,
            "assumptionsExclusions": offer.assumptions_exclusions,
            "fringesPercent": str(offer.fringes_percent),
            "handlingPercent": str(offer.handling_percent),
            "markupPercent": str(offer.markup_percent),
            "productionInsurancePercent": str(offer.production_insurance_percent),
            "productionFeePercent": str(offer.production_fee_percent),
            "postMarkupPercent": str(offer.post_markup_percent),
            "postInsurancePercent": str(offer.post_insurance_percent),
            "postTaxPercent": str(offer.post_tax_percent),
            "customFeeNames": (offer.details or {}).get("customFeeNames", {}),
            "categoryExternalMarkup": (offer.details or {}).get(
                "categoryExternalMarkup", {}
            ),
            "deliverables": [
                {
                    "id": str(x.id),
                    "quantity": x.quantity,
                    "duration": x.duration,
                    "durationUnit": x.duration_unit,
                    "notes": x.notes,
                    "sortOrder": x.sort_order,
                }
                for x in deliverables
            ],
            "scheduleEntries": [
                {
                    "id": str(x.id),
                    "phaseType": x.phase_type,
                    "days": x.days,
                    "hoursPerDay": x.hours_per_day,
                    "notes": x.notes,
                    "sortOrder": x.sort_order,
                }
                for x in schedule_entries
            ],
        },
        "project": {
            "id": str(project.id),
            "name": project.name,
            "agencyName": project.agency_name,
            "clientName": project.client_name
            or (project.client.name if project.client else None),
            "brandName": project.brand_name,
            "clientManagers": client_managers_list,
        },
        "vendor": {
            "name": vendor.name,
            "companyName": company_name or None,
            "logoUrl": logo_url,
        },
        "collaborators": collaborators_list,
        "categories": categories_list,
        "shareToken": share.token if share else None,
    }


@csrf_exempt
@require_http_methods(["GET"])
@require_groups("VENDOR", "SYSTEM")
def offer_export_data(request, offer_id):
    try:
        offer = Offer.objects.select_related(
            "project",
            "project__vendor",
            "project__client",
        ).get(id=offer_id, deleted_at__isnull=True)
    except Offer.DoesNotExist:
        return JsonResponse({"error": "Offer not found"}, status=404)

    user_vendor_id = request.user_data.get("vendor_id")
    if (
        not user_vendor_id
        or not offer.project
        or str(offer.project.vendor_id) != user_vendor_id
    ):
        return JsonResponse({"error": "Access denied"}, status=403)

    return JsonResponse(_build_offer_export_data(offer))


@csrf_exempt
@require_http_methods(["GET"])
@public_endpoint
def share_export_data(request, token):
    try:
        share = Share.objects.select_related(
            "offer",
            "offer__project",
            "offer__project__vendor",
            "offer__project__client",
        ).get(token=token)
    except Share.DoesNotExist:
        return JsonResponse({"error": "Share not found"}, status=404)

    if not share.is_active:
        return JsonResponse({"error": "Share link is no longer active"}, status=403)

    if share.offer.project and share.offer.project.deleted_at is not None:
        return JsonResponse({"error": "Project is archived"}, status=403)

    offer = share.offer
    if offer.deleted_at is not None:
        return JsonResponse({"error": "Offer not found"}, status=404)

    return JsonResponse(_build_offer_export_data(offer))
